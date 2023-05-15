from fastapi import UploadFile, File
import os
from src import cfg
from src.models import ADOC_M, ADOC_HISTORY_M, AUTHOR_M
import openpyxl
import re

from src.utils.mystrings import strip_last, removing_leading_whitespaces, str_cleanup


# import asyncpg


async def adoc_get_all_count():
    content = {"msg": f"Unknown error"}
    # log = set_logger(settings.WELL_FILE_LOG)
    try:
        all_count = await ADOC_M.objects.count()
        # log.info(f"count load successfuly: {all_count}")
        content = {"msg": "Success", "count": all_count}
        return content
    except Exception as e:
        str_err = "Exception occurred " + str(e)
        content = {"msg": f"reload fail. can't read count from table {ADOC_M.Meta.tablename}", "err": str(e)}
        print(str_err)
        # log.info(str_err)
    return content


async def adoc_get_by_id(id_str: str):
    content = {"msg": f"Unknown error"}
    try:
        if id_str.isnumeric() :
            id_int =  int(id_str)
        else:
            id_int = 0
        all_ = await ADOC_M.objects.all(ADOC_M.id == id_int)
        all_count = len(all_)
        content = {
            "msg": "Success",
            "count": all_count,
            "data": all_
        }
        return content
    except Exception as e:
        str_err = "Exception occurred " + str(e)
        content = {"msg": f"reload fail. can't read count from table {ADOC_M.Meta.tablename}", "err": str(e)}
        print(str_err)
        # log.info(str_err)
    return content



async def adoc_upload_file(file: UploadFile = File(...)):
    content = {"msg": f"Unknown error"}
    try:

        if not os.path.isdir(cfg.FOLDER_UPLOAD):
            os.mkdir(cfg.FOLDER_UPLOAD)
        file_in = file.filename
        file_out = file.filename.replace(" ", "-")

        file_name = os.path.join(os.getcwd(), cfg.FOLDER_UPLOAD, file_out)

        if os.path.exists(file_name):
            os.remove(file_name)
            print(f"{file_name} - deleted !")

        contents = file.file.read()
        with open(file_name, 'wb') as f:
            f.write(contents)

        adoc_hist = ADOC_HISTORY_M(file_in=file_in, file_out=file_out, file_out_path=file_name)
        await adoc_hist.upsert()

        all_count = 1
        content = {"msg": "Success", "count": all_count, "filename": file.filename}

        await adoc_excel_file_read(file_name)
        return content
    except Exception as e:
        str_err = "Exception occurred " + str(e)
        content = {"msg": f"Error. There was an error uploading the file", "err": str(e)}
        print(str_err)
        # log.info(str_err)
    finally:
        file.file.close()

    return content


async def adoc_excel_file_read(file_in: str):
    print(f"File input {file_in}")

    await ADOC_M.objects.delete(each=True)
    await AUTHOR_M.objects.delete(each=True)

    # Define variable to load the wookbook
    wookbook = openpyxl.load_workbook(file_in)

    # Define variable to read the active sheet:
    worksheet = wookbook.active

    # Iterate the loop to read the cell values
    # worksheet.max_row
    # worksheet.max_column
    # 18 колонок
    cnt = 0
    try:
        authors = []

        for value in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=18, values_only=True):
            print(f"{cnt}")
            cnt = cnt + 1
            folder_root = ''
            folder_link = ''
            folder_short = ''
            folder_name = ''
            rgf = ''
            tgf_hmao = ''
            tgf_ynao = ''
            tgf_kras = ''
            tgf_ekat = ''
            tgf_omsk = ''
            tgf_novo = ''
            tgf_more = ''
            tgf_tmn = ''
            tgf = ''
            report_name = ''
            author_name = ''
            year_str = ''
            year_int = 0
            territory_name = ''
            comments = ''

            if value[0]:  # Путь полный
                folder_root = str_get_full_path_with_format(str(value[0]))
                folder_link = folder_root  # Гиперссылка
                folder_short = str_get_folder(folder_root)
                folder_name = folder_short
            if value[4]:
                rgf = str_tgf_format(str(value[4]))
            if value[5]:
                tgf_hmao = str_tgf_format(str(value[5]))
            if value[6]:
                tgf_ynao = str_tgf_format(str(value[6]))
            if value[7]:
                tgf_kras = str_tgf_format(str(value[7]))
            if value[8]:
                tgf_ekat = str_tgf_format(str(value[8]))
            if value[9]:
                tgf_omsk = str_tgf_format(str(value[9]))
            if value[10]:
                tgf_novo = str_tgf_format(str(value[10]))
            if value[11]:
                tgf_more = str_tgf_format(str(value[11]))
            if value[12]:
                tgf_tmn = str_tgf_format(str(value[12]))

            if len(tgf_tmn):
                tgf = 'ТюмТГФ'
            if len(tgf_more):
                tgf = 'МорскойТГФ'
            if len(tgf_novo):
                tgf = 'НовосибТГФ'
            if len(tgf_omsk):
                tgf = 'ОмскТГФ'
            if len(tgf_ekat):
                tgf = 'ЕкатерТГФ'
            if len(tgf_kras):
                tgf = 'КраснТГФ'
            if len(tgf_ynao):
                tgf = 'ЯНТГФ'
            if len(tgf_hmao):
                tgf = 'ХМТГФ'
            if len(rgf):
                tgf = 'РГФ'

            # print(tgf)

            if value[14]:  # Отчет
                str_tmp1 = str(value[14]).strip().replace("_x001E_", "")
                report_name = str_cleanup(str_tmp1)

            if value[15]:  # Авторы
                author_name = str(value[15]).strip()
                author_tmp = str_clean(author_name)
                if len(author_tmp) > 2:
                    result = re.match(r'^0-9', author_tmp)
                    result2 = re.match(r'[^\D]', author_tmp)
                    if not result or not result2:
                        authors.append(author_tmp)

            if value[16]:  # Год
                year_str = str(value[16]).strip()
                # print(year_str)
                year_tmp = year_str.split()
                if len(year_tmp):  # разбираем такое 2009   2009   2010
                    year_int = int(year_tmp[0])

            if value[17]:  # Территория
                territory_name = str(value[17]).strip()
                # print(territory_name)

            await ADOC_M(
                folder_root=folder_root,
                folder_link=folder_link,
                folder_short=folder_short,
                folder_name=folder_name,
                rgf=rgf,
                tgf_hmao=tgf_hmao,
                tgf_ynao=tgf_ynao,
                tgf_kras=tgf_kras,
                tgf_ekat=tgf_ekat,
                tgf_omsk=tgf_omsk,
                tgf_novo=tgf_novo,
                tgf_more=tgf_more,
                tgf_tmn=tgf_tmn,
                tgf=tgf,
                report_name=report_name,
                author_name=author_name,
                year_str=year_str,
                year_int=year_int,
                territory_name=territory_name,
                comments=comments
            ).save()

            # str_tmp1 = ''
        authors_str = ",".join(authors)
        authors2 = authors_str.split(",")
        authors_tmp2 = sorted(set(authors2))
        for author in authors_tmp2:
            if len(author) > 2:
                await AUTHOR_M(author_name=author).save()
                # authors.append(author)

    except Exception as e:
        str_err = "Exception occurred " + str(e)
        content = {"msg": f"ERROR!!! cnt is {cnt}", "err": str(e)}
        print(str_err)

    # for i in range(1, 5):
    # for col in worksheet.iter_cols(0, 18):
    # print(col[i].value )
    # print(col[1].value)
    #         # end = "\t\t"
    #     print('')

    # print(worksheet.max_column)
    # print(worksheet.max_row)

def str_tgf_format(str_in: str):
    if len(str_in) > 1 :
        return str_in
    else:
        return ''


def str_get_full_path_with_format(str_in: str):
    if len(str_in) and str_in.startswith('\\'):
        if str_in.endswith('\\'):
            str_tmp = str_in.removesuffix('\\')
            return str_tmp
        else:
            return str_in
    else:
        return ''


def str_get_folder(str_in: str):
    if len(str_in) and str_in.startswith('\\'):
        str_arr = str_in.rsplit('\\', 1)
        if len(str_arr):
            str_tmp = str(str_arr[1])
        else:
            str_tmp = ""
        return str_tmp
    else:
        return ''


def str_clean(str_in: str):
    str_tmp = str_cleanup(str_in)
    str_tmp = removing_leading_whitespaces(str_tmp)
    # str_tmp = cleanupstring(str_tmp)

    str_tmp = str_tmp.lstrip() \
                  .rstrip() \
                  .replace("Авторы:", "") \
                  .replace("Леонов А.П.", "Леонов А.П.,") \
                  .replace("Ляхов С.В.", "Ляхов С.В.,") \
                  .replace("Чунихина Л.Д.", "Чунихина Л.Д.,") \
                  .replace("Херувимова Е.В.", "Херувимова Е.В.,") \
                  .replace("Брадучан Ю.В.", "Брадучан Ю.В.,") \
                  .replace("Дьяконова Ю.А", "Дьяконова Ю.А.,") \
                  .replace("Кос.И.М.", "Кос И.М.,") \
                  .replace("ЗАО НПК «Форум»", " ") \
                  .replace("и др.", ",") \
                  .replace(" др.", ",") \
                  .replace(" др", ",") \
                  .replace("отв. исполнитель", "") \
                  .replace("г.", "") \
                  .replace("0", "") \
                  .replace("1", "") \
                  .replace("2", "") \
                  .replace("3", "") \
                  .replace("4", "") \
                  .replace("5", "") \
                  .replace("6", "") \
                  .replace("7", "") \
                  .replace("8", "") \
                  .replace("9", "") \
                  .replace("0", "") \
                  .replace("/", "") \
                  .replace("\n", "") \
                  .replace("-", "") \
                  .replace(" и ", ",") \
                  .replace("  ", "") \
                  .replace(")", "") \
                  .replace("(", ",") \
                  .replace(" - ", ",") \
                  .replace(" Есть ", ",") \
                  .replace("\xa0", ",") \
                  .replace("_x001E_", "") \
                  .replace("\t", "") \
                  .replace("  ", "") \
                  .replace(", ", ",") \
                  .replace("См_____", " ") \
                  .replace("См____", " ") \
                  .replace(". ", ".") \
                  .lstrip() \
                  .rstrip() \
                  .strip() + ','
    # .replace(". и", ".") \
    # if str_tmp.startswith(" "):
    #     str_tmp = str_tmp.replace(" ", "")
    if str_tmp.endswith(" "):
        str_tmp = strip_last(str_tmp)

    if len(str_tmp):
        result = re.match("\s", str_tmp)
        if result:
            str_tmp = str_tmp.strip()
    return str_tmp


async def adoc_get_all():
    content = {"msg": f"Unknown error"}
    try:
        all_ = await ADOC_M.objects.all()
        all_count = len(all_)
        content = {
            "msg": "Success",
            "count": all_count,
            "data": all_
        }
        return content
    except Exception as e:
        str_err = "Exception occurred " + str(e)
        content = {"msg": f"reload fail. can't read count from table {ADOC_M.Meta.tablename}", "err": str(e)}
        print(str_err)
        # log.info(str_err)
    return content

async def adoc_get_all_with_limit_offset(limit=1, offset=1):
    content = {"msg": f"Unknown error"}
    try:
        all_ = await ADOC_M.objects.offset(offset).limit(limit).all()
        all_count = len(all_)
        content = {
            "msg": "Success",
            "count": all_count,
            "data": all_
        }
        return content
    except Exception as e:
        str_err = "Exception occurred " + str(e)
        content = {"msg": f"reload fail. can't read count from table {ADOC_M.Meta.tablename}", "err": str(e)}
        print(str_err)
        # log.info(st


async def adoc_get_all_by_author(author):
    content = {"msg": f"Unknown error"}
    try:
        all_ = await ADOC_M.objects.filter(ADOC_M.author_name.icontains(author)).all()
        # books = await Book.objects.filter(Book.author.name.icontains("tolkien")).all()
        all_count = len(all_)
        content = {
            "msg": "Success",
            "count": all_count,
            "data": all_
        }
        return content
    except Exception as e:
        str_err = "Exception occurred " + str(e)
        content = {"msg": f"reload fail. can't read count from table {AUTHOR_M.Meta.tablename}", "err": str(e)}
        print(str_err)
        # log.info(str_err)
    return content


async def adoc_get_all_by_report(report):
    content = {"msg": f"Unknown error"}
    try:
        all_ = await ADOC_M.objects.filter(ADOC_M.report_name.icontains(report)).all()
        all_count = len(all_)
        content = {
            "msg": "Success",
            "count": all_count,
            "data": all_
        }
        return content
    except Exception as e:
        str_err = "Exception occurred " + str(e)
        content = {"msg": f"reload fail. can't read count from table {AUTHOR_M.Meta.tablename}", "err": str(e)}
        print(str_err)
        # log.info(str_err)
    return content


async def adoc_get_all_by_year(year):
    content = {"msg": f"Unknown error"}
    try:
        all_ = await ADOC_M.objects.all(ADOC_M.year_int == year)
        all_count = len(all_)
        content = {
            "msg": "Success",
            "count": all_count,
            "data": all_
        }
        return content
    except Exception as e:
        str_err = "Exception occurred " + str(e)
        content = {"msg": f"reload fail. can't read count from table {AUTHOR_M.Meta.tablename}", "err": str(e)}
        print(str_err)
        # log.info(str_err)
    return content


async def adoc_get_all_by_yeargte(year):
    content = {"msg": f"Unknown error"}
    try:
        # year_int__gte = 0, year_int__gte = year
        all_ = await ADOC_M.objects.filter(
            (ADOC_M.year_int > 0) & (ADOC_M.year_int >= year)
        ).all()
        all_count = len(all_)
        content = {
            "msg": "Success",
            "count": all_count,
            "data": all_
        }
        return content
    except Exception as e:
        str_err = "Exception occurred " + str(e)
        content = {"msg": f"reload fail. can't read count from table {AUTHOR_M.Meta.tablename}", "err": str(e)}
        print(str_err)
        # log.info(str_err)
    return content


async def adoc_get_all_by_yearlte(year):
    content = {"msg": f"Unknown error"}
    try:
        all_ = await ADOC_M.objects.filter(
            (ADOC_M.year_int > 0) & (ADOC_M.year_int <= year)
        ).all()
        all_count = len(all_)
        content = {
            "msg": "Success",
            "count": all_count,
            "data": all_
        }
        return content
    except Exception as e:
        str_err = "Exception occurred " + str(e)
        content = {"msg": f"reload fail. can't read count from table {AUTHOR_M.Meta.tablename}", "err": str(e)}
        print(str_err)
        # log.info(str_err)
    return content


async def adoc_get_all_by_year_between(year1, year2):
    content = {"msg": f"Unknown error"}
    try:
        all_ = await ADOC_M.objects.filter(
            (ADOC_M.year_int >= year1) & (ADOC_M.year_int <= year2)
        ).all()
        all_count = len(all_)
        content = {
            "msg": "Success",
            "count": all_count,
            "data": all_
        }
        return content
    except Exception as e:
        str_err = "Exception occurred " + str(e)
        content = {"msg": f"reload fail. can't read count from table {AUTHOR_M.Meta.tablename}", "err": str(e)}
        print(str_err)
        # log.info(str_err)
    return content


async def adoc_get_all_years():
    content = {"msg": f"Unknown error"}
    try:
        years = []
        years_uniq = []
        all_ = await ADOC_M.objects.all()
        for year in all_:
            if year.year_int > 0 :
                years.append(year.year_int)
        years_uniq = sorted(set(years))
        print(years_uniq)
        all_count = len(years_uniq)
        content = {
            "msg": "Success",
            "count": all_count,
            "data": years_uniq
        }
        return content
    except Exception as e:
        str_err = "Exception occurred " + str(e)
        content = {"msg": f"reload fail. can't read count from table {AUTHOR_M.Meta.tablename}", "err": str(e)}
        print(str_err)
        # log.info(str_err)
    return content


# async def files_get_by_root_folder(root_folder: str):
#     content = {"msg": f"Unknown error"}
#     # log = set_logger(settings.WELL_FILE_LOG)
#     try:
#         all_ = await ADOC_M.objects.all(ADOC_M.root_folder == root_folder)
#         all_count = len(all_)
#         # log.info(f"count load successfuly: {all_count}")
#         content = {
#             "msg": "Success",
#             "count": all_count,
#             "data": all_
#         }
#         return content
#     except Exception as e:
#         str_err = "Exception occurred " + str(e)
#         content = {"msg": f"reload fail. can't read count from table {ADOC_M.Meta.tablename}", "err": str(e)}
#         print(str_err)
#         # log.info(str_err)
#     return content


# async def files_get_by_query(str_query: str):
#     content = {"msg": f"Unknown error"}
#     # log = set_logger(settings.WELL_FILE_LOG)
#     try:
#         # books = (
#         #     await Book.objects.filter(Book.author.name.icontains("tolkien"))
#         #     .order_by(Book.year.desc())
#         #     .all()
#         # )
#         # надо сохранять поисковый запрос...
#         print(str_query)
#         # order_by(Book.year.desc())
#         all_ = await FILES_M.objects.filter(FILES_M.file_path.icontains(str_query)).all()
#         all_count = len(all_)
#         # log.info(f"count load successfuly: {all_count}")
#         content = {
#             "msg": "Success",
#             "count": all_count,
#             "data": all_
#             }
#         return content
#     except Exception as e:
#         str_err = "Exception occurred " + str(e)
#         content = {"msg": f"reload fail. can't read from table {FILES_M.Meta.tablename}", "err": str(e)}
#         print(str_err)
#         # log.info(str_err)
#     return content
